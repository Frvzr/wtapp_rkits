import os
import pandas as pd
from math import floor
import logging
import pathlib
from openpyxl.styles import Alignment
import decimal

DIR = os.getcwd()
FILE_PATH = f'{DIR}\\required_redress_kit.xlsx'

_log_format = f"%(asctime)s - [%(levelname)s] - %(name)s - (%(filename)s).%(funcName)s(%(lineno)d) - %(message)s"

def get_file_handler():
    """_summary_

    Returns:
        _type_: _description_
    """
    file_handler = logging.FileHandler("log.log")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(_log_format))
    return file_handler

def get_stream_handler():
    """_summary_

    Returns:
        _type_: _description_
    """
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(logging.Formatter(_log_format))
    return stream_handler

def get_logger(name):
    """_summary_

    Args:
        name (_str_): _description_

    Returns:
        _type_: _description_
    """
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    logger.addHandler(get_file_handler())
    logger.addHandler(get_stream_handler())
    return logger


logger = get_logger(__name__)


def get_data_from_excel(FILE_PATH):
    """
    Функция получает данные с файла

    Args:
        FILE_PATH (_str_): Путь к файлу с обрабатываемыми данными

    Returns:
        (_dict_): Функция возвращает данные с 3-х заранее известных страниц файла
    """
    try:
        logger.info('Start programm')
        redress = pd.read_excel(FILE_PATH, sheet_name='Required redress kits')
        required_redress_kits = {"series": []}
        for redress_kit, v in redress.groupby("Redress kit", sort=False):
            required_redress_kits["series"].append({"redress_kit": redress_kit.upper(), "total": []})
            for quantity_on_store, required_quantity in zip(v["Q-ty on store"], v["Req qty"]):
                required_redress_kits["series"][-1]["total"].append({"q-ty on store": quantity_on_store, "required": required_quantity})

        rk_bom = pd.read_excel(FILE_PATH, sheet_name='Redress kit BOM')
        redress_kit_bom = {"series": []}
        for redress_kit_, redress_kit_items in rk_bom.groupby("Redress Part Number"):
            redress_kit_bom["series"].append({"redress kit": redress_kit_.upper(), "consist": []})
            for w, s, t in zip(redress_kit_items["Item Part Number"], redress_kit_items["Quantity pr."], redress_kit_items['Description']):
                redress_kit_bom["series"][-1]["consist"].append({'item': w, 'description': t, 'qty': s})
        
        qty_on_store = pd.read_excel(FILE_PATH, sheet_name='Pivot Stock')
        qty_on_store_data = dict(zip(qty_on_store['Part Number'], qty_on_store['Sum of QTY']))
        qty_on_store_data_upper = {k.upper(): v for k, v in qty_on_store_data.items()}

        return required_redress_kits, redress_kit_bom, qty_on_store_data_upper
    except Exception as e:
        logger.critical(e)
        
        
def merge_consist(required_redress_kits, redress_kit_bom):
    """_summary_

    Args:
        need_redress_kits (_dict_): данные с необходимыми наборами зип
        redress_kit_bom (_dict_): данные с составляющими запчастями, входящие в набор зип

    Returns:
        (_dict_): данные с необходимым кол-вом наборов зип и их составом
    """
    try:
        required_with_items = {'series': []}
        for a, b in required_redress_kits.items():
            for z in b:
                for k, v in redress_kit_bom.items():
                    for i in v:
                        if i["redress kit"] == z['redress_kit']:
                            required_with_items['series'].append({'redress_kit':i['redress kit'], "total": z["total"], "consist": i["consist"]})
        return required_with_items
    except Exception as e:
        logger.critical(e)


def merge_store(qty_on_store_data, required_with_items):
    """_summary_

    Args:
        qty_on_store_data (dict): словарь с количеством зип на складе
        required_with_items (dict): словарь с необходимым кол-вом наборов и их составом

    Returns:
        dict: словарь с максимальным кол-вом наборов, которые мы можем собрать
    """
    try:
        max_collect_redress = {'series': []}
        for a, b in required_with_items.items():
            for i in b:
                required = i['total'][-1]['required']
                qty_on_store = {'qty_on_store': []}
                max_collect_items = {'max_collect_items': []}
                reserved = {'reserved': []}
                for y in i['consist']:
                    if y['item'] not in qty_on_store_data:
                        print(f"This item: {y['item']} is out of stock")
                        qty_on_store_data[y['item']] = 0
                    for k, v in qty_on_store_data.items():
                        if y['item'] == k.upper():
                            max_collect_item = floor(v / y['qty'])
                            max_collect_items['max_collect_items'].append({"item": k.upper(), "qty": max_collect_item})
                            qty_on_store['qty_on_store'].append({'item': k, 'qty': v})

                            if not pd.isna(required):
                                reserved = get_reserved(required, i, qty_on_store_data)
                        
                res = get_min_data(max_collect_items)
                if not pd.isna(required) and res > required:
                    res = required
                if pd.isna(required):
                    reserved = get_reserved(res, i, qty_on_store_data)
                qty_on_store_data = update_store(qty_on_store_data, reserved['reserved'])
                
                max_collect_redress['series'].append({'redress_kit':i['redress_kit'], "total": i["total"], "consist": i["consist"], "max_collect_items": max_collect_items["max_collect_items"], "maximum_collect": res, 'qty_on_store': qty_on_store['qty_on_store'], 'reserved': reserved['reserved']})
        return max_collect_redress, qty_on_store_data
    except Exception as e:
        logger.critical(e)
        
def get_reserved(res, redress, qty_on_store_data):
    reserved = {'reserved': []}
    required = res if res >= 0 else 1
    for y in redress['consist']:
        for k, v in qty_on_store_data.items():
            if y['item'] == k.upper():
                if v > 0:
                    req = y['qty'] * int(required)
                    reserv = v - req
                    if reserv >= 0:
                        reserved['reserved'].append({"item": k.upper(), "qty": req})
                    elif reserv < 0:
                        reserved['reserved'].append({"item": k.upper(), "qty": v})
                else:
                    reserved['reserved'].append({"item": k.upper(), "qty": 0})
    return reserved

def get_min_data(data):
    try:
        min_data = []
        if data['max_collect_items']:
            for k, v in data.items():
                for i in v:
                    min_data.append(i['qty'])
            return min(min_data)
        else:
            return 0
    except Exception as e:
        logger.error(e)


def update_store(qty_on_store_data, reserved):
    try:         
        for j in reserved:
            for k, v in qty_on_store_data.items():  
                if j['item'] == k.upper() and (v - j['qty']) > 0:
                    qty_on_store_data[k] = v - j['qty']                             
                elif j['item'] == k.upper() and (v - j['qty']) <= 0:
                    qty_on_store_data[k] = 0
        return qty_on_store_data
    except Exception as e:
        logger.error(e)


def handling_data(data, updated_store):
    out_data = {'Redress Kit':[],
                'Qty on store': [],
                'Required': [],
                'Can collect': [],
                'Item': [],
                'Qty per kit':[],
                'Description': [],
                'Need to order': [],
                'Reserved':[]
    }
    
    store_data = {'Item': [],
                  'Quantity': []}
    
    try:
        for i in data['series']:
            required = i['total'][-1]['required']
            max_collect = i['maximum_collect']
            
            if pd.isna(required) and max_collect == 0:
                 required = 1
            elif pd.isna(required) and max_collect > 0:
                required = max_collect
            for j in i['consist']:
                need_qty = j['qty'] * required
                for a in i['qty_on_store']:
                    for b in i['reserved']:
                        if j['item'].upper() == a['item'].upper() and j['item'].upper() == b['item'].upper():
                            out_data["Redress Kit"].append(i['redress_kit'])
                            out_data['Qty on store'].append(i['total'][-1]['q-ty on store'])
                            out_data['Required'].append(required)
                            out_data['Can collect'].append(max_collect)
                            out_data['Item'].append(a['item'])
                            out_data['Qty per kit'].append(j['qty'])
                            out_data['Description'].append(j['description'])
                            if (need_qty - a['qty']) > 0:
                                out_data['Need to order'].append(need_qty - a['qty'])
                            else:
                                out_data['Need to order'].append(0)
                            out_data['Reserved'].append(b['qty']) 

        for k, v in updated_store.items():
            store_data['Item'].append(k)
            store_data['Quantity'].append(v)              
        
        return out_data, store_data
    except Exception as e:
        logger.error(e)

def bg_header(x):
    return "background-color: #9ccc65"

def get_text_color(val):
    color = 'red' if val == 'Need to order' else 'black'
    return 'color: %s' % color

def get_center_text(val):
    return 'text-align: center'
        
def output_data(all_data, store_data):
    try:
        out_path = pathlib.Path('can_collect_redress_kits.xlsx')
        with pd.ExcelWriter('can_collect_redress_kits.xlsx', engine="openpyxl", mode='a', if_sheet_exists="replace") if out_path.exists() else pd.ExcelWriter('can_collect_redress_kits.xlsx', engine="openpyxl", mode='w') as wb:
            df = pd.DataFrame(all_data)
            df_store = pd.DataFrame(store_data)
            SHEETNAME = 'Collect Redress Kit'
            SHEETNAME_STORE = 'Store'
  
            df.style.applymap_index(bg_header, axis=1).applymap_index(get_text_color, axis=1).applymap_index(get_center_text, axis=1).set_properties(**{'text-align': 'center'}).to_excel(wb, sheet_name=SHEETNAME, index=False)
            df_store.to_excel(wb, sheet_name=SHEETNAME_STORE, index=False)
            df_store.style.applymap_index(bg_header, axis=1).applymap_index(get_text_color, axis=1).applymap_index(get_center_text, axis=1).set_properties(**{'text-align': 'center'}).to_excel(wb, sheet_name=SHEETNAME_STORE, index=False)
            
            ws = wb.sheets[SHEETNAME]
            ws.auto_filter.ref='a:g'
            ws.column_dimensions['A'].width = 13
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['i'].width = 10
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['E1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['F1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['G1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['H1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['I1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws_store = wb.sheets[SHEETNAME_STORE]
            ws_store.auto_filter.ref='a:b'
            ws_store.column_dimensions['A'].width = 20
            ws_store.column_dimensions['B'].width = 11
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            logger.info('End programm')
    except Exception as e:
        logger.error(e)
            
            
def main():
    required_redress_kits, redress_kit_bom, qty_on_store_data = get_data_from_excel(FILE_PATH)
    required_with_items = merge_consist(required_redress_kits, redress_kit_bom)
    raw_data, updated_store = merge_store(qty_on_store_data, required_with_items)
    #print(raw_data)
    all_data, store_data = handling_data(raw_data, updated_store)
    output_data(all_data, store_data)
    
if __name__ == '__main__':
    main()
    