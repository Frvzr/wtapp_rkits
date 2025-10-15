from constants import (
    INPUT_FILE, FILE_PATH, SHEETNAME_OUTPUT, SHEETNAME_STORE_OUTPUT,
    SHEETNAME_INPUT_BOM, SHEETNAME_INPUT_STOCK, SHEETNAME_INPUT_REQUIRED,
    OUTPUT_FILE
)
import pandas as pd
from math import floor
from pathlib import Path
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from logger import get_logger

from typing import Dict, List, Tuple, Union

logger = get_logger(__name__)


def get_data_from_excel_stock(path: str) -> Tuple[Dict, Dict, Dict]:
    """
    Получение данных о запасах из Excel файла

    Args:
        path: Путь к файлу Excel

    Returns:
        Кортеж с данными:
        - items_by_stock: Данные по складам
        - total: Общие количества
        - actual_serial_items: Серийные номера
    """
    try:
        logger.info('Starting program')

        # Чтение данных с явным указанием engine
        stock_dataframe = pd.read_excel(
            path,
            sheet_name='Stock 2023',
            engine='openpyxl'
        )

        # Проверка наличия необходимых колонок
        required_columns = ['Part Number', 'Stock', 'QTY', 'SN']
        missing_cols = [col for col in required_columns if col not in stock_dataframe.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")

        # Группировка данных - исправленная версия
        grouped_data = stock_dataframe.groupby(['Part Number', 'Stock'])['QTY'].sum()

        items_by_stock = {'Main': {}, 'Ru Ops': {}}
        for (part_num, stock), qty in grouped_data.items():
            if isinstance(stock, str) and stock.lower() == 'main':
                items_by_stock['Main'][part_num] = qty
            else:
                items_by_stock['Ru Ops'][part_num] = qty

        # Расчет общего количества
        all_part_numbers = set(items_by_stock['Main'].keys()).union(set(items_by_stock['Ru Ops'].keys()))
        total = {
            k: items_by_stock['Main'].get(k, 0) + items_by_stock['Ru Ops'].get(k, 0)
            for k in all_part_numbers
        }


        # Обработка серийных номеров
        serial_grouped = stock_dataframe.groupby(['Part Number', 'Stock', 'SN'])['QTY'].sum()
        actual_serial_items = {
            item_sn: qty for item_sn, qty in serial_grouped.items() if qty != 0
        }

        logger.info(f"Data from {SHEETNAME_INPUT_STOCK} uploaded successfully")

        return items_by_stock, total, actual_serial_items

    except ValueError as ve:
        logger.critical(f"Value error in {INPUT_FILE}: {str(ve)}", exc_info=True)
        raise
    except KeyError as ke:
        logger.critical(f"Column error in {INPUT_FILE}: {str(ke)}", exc_info=True)
        raise
    except Exception as e:
        logger.critical(f"Unexpected error processing {INPUT_FILE}: {str(e)}", exc_info=True)
        raise


def get_data_from_excel_required_redress(path: str) -> Dict:
    """Получение данных о требуемых наборах из Excel"""
    try:
        redress = pd.read_excel(path, sheet_name=SHEETNAME_INPUT_REQUIRED, engine='openpyxl')
        required_redress_kits = {"Required redress kit": []}

        for redress_kit, req_qty in redress.groupby("Redress kit", sort=False):
            kit_data = {
                "redress_kit": redress_kit.upper(),
                "total": [
                    {
                        "q-ty on store": row["Q-ty on store"],
                        "required": row["Req qty"]
                    }
                    for _, row in req_qty.iterrows()
                ]
            }
            required_redress_kits["Required redress kit"].append(kit_data)

        logger.info(f"Data from {SHEETNAME_INPUT_REQUIRED} uploaded successfully")
        return required_redress_kits

    except Exception as e:
        logger.critical(f"Error reading required redress data: {str(e)}", exc_info=True)
        raise


def get_data_from_excel_redress_kits_bom(path: str) -> Dict:
    """Получение данных BOM из Excel"""
    try:
        rk_bom = pd.read_excel(path, sheet_name=SHEETNAME_INPUT_BOM, engine='openpyxl')
        redress_kit_bom = {"redress kit consist": []}

        for redress_kit_, redress_kit_items in rk_bom.groupby("Redress Part Number"):
            kit_data = {
                "redress kit": redress_kit_.upper(),
                "consist": [
                    {
                        'item': row["Item Part Number"],
                        'description': row['Description'],
                        'qty': row["Quantity pr."]
                    }
                    for _, row in redress_kit_items.iterrows()
                    if row["Quantity pr."] not in (0, '0')
                ]
            }
            redress_kit_bom["redress kit consist"].append(kit_data)

        logger.info(f"Data from {SHEETNAME_INPUT_BOM} uploaded successfully")
        return redress_kit_bom

    except Exception as e:
        logger.critical(f"Error reading BOM data: {str(e)}", exc_info=True)
        raise


def merge_consist(required_redress_kits: Dict, redress_kit_bom: Dict) -> Dict:
    """Объединение данных о требуемых наборах и BOM"""
    required_with_items = {'Items for redress kits': []}

    for req_kit in required_redress_kits['Required redress kit']:
        for consist in redress_kit_bom["redress kit consist"]:
            if consist["redress kit"] == req_kit['redress_kit']:
                required_with_items['Items for redress kits'].append({
                    'redress_kit': consist['redress kit'],
                    "total": req_kit["total"],
                    "consist": consist["consist"]
                })

    logger.info("Merging required redress kits and BOM completed")
    return required_with_items


def merge_store(qty_on_store_data: Dict, required_with_items: Dict, serial_items: Dict) -> Tuple[Dict, Dict]:
    """Объединение данных со складом"""
    max_collect_redress = {'maximum collect rkits': []}
    for item in required_with_items['Items for redress kits']:
        required = item['total'][-1]['required']
        qty_on_store = {'qty_on_store': []}
        max_collect_items = {'max_collect_items': []}
        reserved = {'reserved': []}
        serial = {'serial_items': []}

        for y in item['consist']:
            item_key = y['item'].upper()

            # Проверка наличия товара на складе
            if item_key not in qty_on_store_data:
                logger.warning(f"Item {item_key} is out of stock")
                qty_on_store_data[item_key] = 0

            # Расчет максимального количества для сбора
            max_collect_item = floor(qty_on_store_data.get(item_key, 0) / y['qty'])
            max_collect_items['max_collect_items'].append({
                "item": item_key,
                "qty": max_collect_item
            })

            qty_on_store['qty_on_store'].append({
                'item': item_key,
                'qty': qty_on_store_data.get(item_key, 0)
            })

            # Обработка серийных номеров
            for sn_item, sn_qty in serial_items.items():
                if sn_item[0].upper() == item_key:
                    serial['serial_items'].append({
                        'sn_item': sn_item[0],
                        'store': sn_item[1],
                        'serial_number': sn_item[2],
                        'sn_qty': sn_qty
                    })

        # Расчет минимального количества
        res = get_min_data(max_collect_items)
        print(res)
        if not pd.isna(required) and res > required:
            res = required

        # Резервирование товаров
        reserved = get_reserved(res, item, qty_on_store_data)

        # Обновление данных склада
        qty_on_store_data = update_store(qty_on_store_data, reserved['reserved'])

        # Формирование итоговых данных
        max_collect_redress['maximum collect rkits'].append({
            'redress_kit': item['redress_kit'],
            "total": item["total"],
            "consist": item["consist"],
            "max_collect_items": max_collect_items["max_collect_items"],
            "maximum_collect": res,
            'qty_on_store': qty_on_store['qty_on_store'],
            'reserved': reserved['reserved'],
            'serial': serial['serial_items']
        })

    logger.info("Merging data with stock completed")
    print(max_collect_redress)
    return max_collect_redress, qty_on_store_data


def get_reserved(res: int, redress: Dict, qty_on_store_data: Dict) -> Dict:
    """Расчет резервирования товаров"""
    reserved = {'reserved': []}
    required = res if res >= 0 else 1

    for item in redress['consist']:
        item_from_consist = item['item'].upper()
        qty_from_consist = item['qty']
        qty_from_store = qty_on_store_data.get(item_from_consist, 0)

        if qty_from_store > 0:
            req = qty_from_consist * int(required)
            reserv = qty_from_store - req
            reserved_qty = req if reserv >= 0 else qty_from_store
        else:
            reserved_qty = 0

        reserved['reserved'].append({
            "item": item_from_consist,
            "qty": reserved_qty
        })

    return reserved


def get_min_data(data: Dict) -> int:
    """Получение минимального значения из данных"""
    min_data = [i['qty'] for i in data['max_collect_items']] if data['max_collect_items'] else [0]
    return min(min_data)


def update_store(qty_on_store_data: Dict, reserved: List[Dict]) -> Dict:
    """Обновление данных склада после резервирования"""
    for j in reserved:
        item_key = j['item'].upper()
        if item_key in qty_on_store_data:
            new_qty = qty_on_store_data[item_key] - j['qty']
            qty_on_store_data[item_key] = max(new_qty, 0)
    return qty_on_store_data


def data_handling_items(data: Dict, items_by_stock: Dict) -> Dict:
    """Формирование итоговых данных для вывода"""
    out_data = {
        'Redress Kit': [],
        'Qty on store': [],
        'Required': [],
        'Can collect': [],
        'Item': [],
        'Qty per kit': [],
        'Description': [],
        'Need to order': [],
        'Reserved': [],
        'Serial Number': [],
        'Main': [],
        'Ru Ops': [],
    }

    for i in data['maximum collect rkits']:
        required = i['total'][-1]['required']
        max_collect = i['maximum_collect']

        # Определение требуемого количества
        if pd.isna(required):
            required = max_collect if max_collect > 0 else 1

        for j in i['consist']:
            req_item = j['item'].upper()
            need_qty = j['qty'] * required
            item_from_main = items_by_stock['Main'].get(req_item, 0)
            item_from_ruops = items_by_stock['Ru Ops'].get(req_item, 0)

            for a in i['qty_on_store']:
                for b in i['reserved']:
                    if req_item == a['item'].upper() and req_item == b['item'].upper():
                        # Обработка серийных номеров
                        serial_found = False
                        for q in i['serial']:
                            if req_item == q['sn_item'].upper():
                                out_data['Serial Number'].append(q['serial_number'])
                                if q['store'] == 'Main':
                                    out_data['Main'].append(q['sn_qty'])
                                    out_data['Ru Ops'].append(0)
                                else:
                                    out_data['Ru Ops'].append(q['sn_qty'])
                                    out_data['Main'].append(0)
                                serial_found = True
                                break

                        if not serial_found:
                            out_data['Serial Number'].append('N/A')
                            out_data['Main'].append(item_from_main)
                            out_data['Ru Ops'].append(item_from_ruops)

                        # Заполнение остальных полей
                        out_data["Redress Kit"].append(i['redress_kit'])
                        out_data['Qty on store'].append(i['total'][-1]['q-ty on store'])
                        out_data['Required'].append(required)
                        out_data['Can collect'].append(max_collect)
                        out_data['Item'].append(a['item'])
                        out_data['Qty per kit'].append(j['qty'])
                        out_data['Description'].append(j['description'])
                        out_data['Need to order'].append(max(need_qty - a['qty'], 0))
                        out_data['Reserved'].append(b['qty'])

    logger.info("Data collection completed")
    return out_data


def data_handling_store(updated_store: Dict) -> Dict:
    """Формирование данных по складу для вывода"""
    store_data = {'Item': [], 'Quantity': []}
    for item_pn, qty in updated_store.items():
        store_data['Item'].append(item_pn)
        store_data['Quantity'].append(qty)

    logger.info("Stock data updated")
    return store_data


def apply_styles(df: pd.DataFrame) -> pd.DataFrame:
    """Применение стилей к DataFrame"""
    return df.style.applymap_index(lambda x: "background-color: #9ccc65", axis=1).applymap_index(lambda x: 'color: red' if x == 'Need to order' else 'color: black', axis=1).applymap_index(lambda x: 'text-align: center', axis=1).set_properties(**{'text-align': 'center'})



def format_sheets(col_dims: Dict, worksheet) -> None:
    """Форматирование листа Excel"""
    for col_name, col_width in col_dims.items():
        worksheet.column_dimensions[col_name].width = col_width
        cell = worksheet[f"{col_name}1"]
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def output_data(all_data: Dict, store_data: Dict) -> None:
    """Сохранение данных в Excel файл с гарантией видимости листов"""
    try:
        # Установка jinja2, если не установлен
        try:
            import jinja2
        except ImportError:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "jinja2"])
            import jinja2

        out_path = Path(OUTPUT_FILE)

        # Создаем DataFrame
        df = pd.DataFrame(all_data)
        df_store = pd.DataFrame(store_data)

        # Определяем режим работы с файлом
        if out_path.exists():
            logger.info(f"Updating existing file: {OUTPUT_FILE}")
            mode = 'a'
            if_sheet_exists = "replace"

            # Загружаем существующую книгу для проверки листов
            book = load_workbook(OUTPUT_FILE)
            # Делаем все существующие листы видимыми
            for sheet in book.worksheets:
                sheet.sheet_state = 'visible'
        else:
            logger.info(f"Creating new file: {OUTPUT_FILE}")
            mode = 'w'
            if_sheet_exists = None

        # Параметры для ExcelWriter
        writer_args = {
            'engine': 'openpyxl',
            'mode': mode,
        }
        if mode == 'a':
            writer_args['if_sheet_exists'] = if_sheet_exists

        with pd.ExcelWriter(OUTPUT_FILE, **writer_args) as writer:
            # Сохраняем данные без стилей, если jinja2 не доступен
            try:
                # Основной лист с стилями
                styled_df = df.style \
                    .applymap_index(lambda x: "background-color: #9ccc65", axis=1) \
                    .applymap_index(lambda x: 'color: red' if x == 'Need to order' else 'color: black', axis=1) \
                    .applymap_index(lambda x: 'text-align: center', axis=1) \
                    .set_properties(**{'text-align': 'center'})

                styled_df.to_excel(writer, sheet_name=SHEETNAME_OUTPUT, index=False)

                # Лист склада с стилями
                styled_store = df_store.style \
                    .applymap_index(lambda x: "background-color: #9ccc65", axis=1) \
                    .applymap_index(lambda x: 'color: red' if x == 'Need to order' else 'color: black', axis=1) \
                    .applymap_index(lambda x: 'text-align: center', axis=1) \
                    .set_properties(**{'text-align': 'center'})

                styled_store.to_excel(writer, sheet_name=SHEETNAME_STORE_OUTPUT, index=False)
            except AttributeError:
                # Если стили не доступны, сохраняем без них
                logger.warning("Jinja2 not available, saving without styles")
                df.to_excel(writer, sheet_name=SHEETNAME_OUTPUT, index=False)
                df_store.to_excel(writer, sheet_name=SHEETNAME_STORE_OUTPUT, index=False)

            # Получаем workbook для дополнительных настроек
            workbook = writer.book

            # Гарантируем видимость листов
            visible_sheets = 0
            for sheetname in [SHEETNAME_OUTPUT, SHEETNAME_STORE_OUTPUT]:
                if sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    ws.sheet_state = 'visible'
                    visible_sheets += 1

            # Если нет видимых листов, делаем первый активный лист видимым
            if visible_sheets == 0 and len(workbook.worksheets) > 0:
                workbook.active.sheet_state = 'visible'
                logger.warning("No visible sheets found, making first sheet visible")

            # Форматирование колонок
            if SHEETNAME_OUTPUT in workbook.sheetnames:
                ws = workbook[SHEETNAME_OUTPUT]
                ws.auto_filter.ref = ws.dimensions
                col_dims_items = {
                    'A': 13, 'B': 8, 'C': 10, 'D': 10, 'E': 12,
                    'F': 8, 'G': 30, 'H': 10, 'I': 10, 'J': 10,
                    'K': 10, 'L': 10
                }
                format_sheets(col_dims_items, ws)

            if SHEETNAME_STORE_OUTPUT in workbook.sheetnames:
                ws_store = workbook[SHEETNAME_STORE_OUTPUT]
                ws_store.auto_filter.ref = ws_store.dimensions
                col_dims_store = {'A': 25, 'B': 11}
                format_sheets(col_dims_store, ws_store)

        logger.info(f"Файл {OUTPUT_FILE} успешно сохранен")

    except PermissionError:
        logger.error(f"Закройте файл перед записью: {OUTPUT_FILE}")
        raise
    except Exception as e:
        logger.critical(f"Ошибка при сохранении в Excel: {str(e)}", exc_info=True)
        raise


def main():
    """Основная функция программы"""
    try:
        logger.info("Starting data processing")

        # Получение данных
        items_by_stock, total, serial_items = get_data_from_excel_stock(FILE_PATH)
        required_redress_kits = get_data_from_excel_required_redress(FILE_PATH)
        redress_kit_bom = get_data_from_excel_redress_kits_bom(FILE_PATH)

        # Обработка данных
        required_with_items = merge_consist(required_redress_kits, redress_kit_bom)
        data, updated_store = merge_store(total, required_with_items, serial_items)
        all_data = data_handling_items(data, items_by_stock)
        store_data = data_handling_store(updated_store)

        # Сохранение результатов
        output_data(all_data, store_data)

        logger.info("Program completed successfully")

    except FileNotFoundError as fnf:
        logger.critical(f"File not found: {INPUT_FILE}")
    except Exception as e:
        logger.critical(f"Program failed: {str(e)}", exc_info=True)
        raise


if __name__ == '__main__':
    main()